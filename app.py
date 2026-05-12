import streamlit as st
import pandas as pd
import re
from io import BytesIO, StringIO
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
 
st.set_page_config(page_title="유창강건 마감 킬러", layout="wide")
st.title("📊 유창강건 세금계산서 누락 체크기 v2")
st.info("(매출대장 + 일반경비) − (카드매출 + 현금영수증 + 선매출 + 선발행) = 세금계산서 발행 대상")
 
# ── 공통 함수 ────────────────────────────────────────────────
def clean(x):
    """거래처명 정규화 (강화판)"""
    x = str(x).strip().strip('"')
    x = re.sub(r'[■□▣▪▫▲△▼▽●○◆◇★☆▶◀♠♣♥♦]', '', x)
    x = x.replace("(주)", "").replace("(유)", "").replace("(株)", "").replace("㈜", "")
    x = x.replace("주식회사", "")
    # 괄호 안 내용 제거 (담당자명, 영문병기 등)
    x = re.sub(r'\([^)]*\)', '', x)
    x = re.sub(r'\[[^\]]*\]', '', x)
    # 공백 제거 전: "유창강건 대동공장" 같은 복합 사업장명도 제거
    x = re.sub(r'\s+\S+?(공장|영업소|지점|지사|사무소|물류센터)\s*$', '', x)
    x = re.sub(r'\s+', '', x)
    x = re.sub(r'[·•・]', '', x)
    # 공백 제거 후에도 끝에 붙는 단순 꼬리
    x = re.sub(r'(서울영업소|부산영업소|대구영업소|광주영업소|영업소|지점|지사|본사|본점)$', '', x)
    # 한글 들어있으면 끝 단일 영문 1자 제거 (욜로가이즈a)
    if re.search(r'[가-힣]', x):
        x = re.sub(r'[a-zA-Z]$', '', x)
    return x.strip().lower()
 
def is_number(s):
    s = str(s).strip().strip('"').replace(',','').replace('(','').replace(')','')
    try: float(s); return True
    except: return False
 
def to_num(x):
    try: return int(float(str(x).replace(',','').replace(' ','')))
    except: return 0
 
def fmt(n):
    try: return f"{int(n):,}"
    except: return "-"
 
def similarity(a, b):
    return SequenceMatcher(None, a, b).ratio()
 
def find_similar(name_clean, tax_raw_list, threshold=0.72):
    results = []
    for row_idx, raw_name in tax_raw_list:
        score = similarity(name_clean, clean(raw_name))
        if score >= threshold and name_clean != clean(raw_name):
            results.append((raw_name, row_idx, score))
    results.sort(key=lambda x: -x[2])
    return results[:3]
 
def read_tsv_cp949(b):
    text = b.decode('cp949', errors='replace')
    return pd.read_csv(StringIO(text), sep='\t', header=0, on_bad_lines='skip')
 
def read_pre_tsv(b):
    """선발행/선매출 (헤더 1행 = 2번째 행)"""
    text = b.decode('cp949', errors='replace')
    df = pd.read_csv(StringIO(text), sep='\t', quotechar='"', header=1, on_bad_lines='skip')
    df = df[df['거래처명'].notna() & (df['거래처명'].astype(str).str.strip() != '')]
    return df
 
# ── session_state ────────────────────────────────────────────
for k in ['매출대장','일반경비','카드매출','현금영수증','계산서','선매출','선발행']:
    if k not in st.session_state:
        st.session_state[k] = None
 
# ── 파일 업로드 UI ──────────────────────────────────────────
st.markdown("### 📁 파일 업로드")
c1, c2 = st.columns(2)
with c1:
    st.markdown("**발행 대상 ➕**")
    col1, col2 = st.columns(2)
    with col1:
        st.caption("1️⃣ 매출대장")
        f1 = st.file_uploader("매출대장", type=['xls','xlsx','csv','txt'], label_visibility="collapsed")
    with col2:
        st.caption("2️⃣ 일반경비")
        f2 = st.file_uploader("일반경비", type=['xls','xlsx','csv'], label_visibility="collapsed")
with c2:
    st.markdown("**제외 대상 ➖**")
    col3, col4 = st.columns(2)
    with col3:
        st.caption("3️⃣ 카드매출")
        f3 = st.file_uploader("카드매출", type=['xls','xlsx','csv'], label_visibility="collapsed")
    with col4:
        st.caption("4️⃣ 현금영수증")
        f4 = st.file_uploader("현금영수증", type=['xls','xlsx','csv'], label_visibility="collapsed")
 
# 신규: 선매출/선발행
st.markdown("**제외 대상 (선처리분) ➖**")
c5a, c5b = st.columns(2)
with c5a:
    st.caption("6️⃣ 선매출 (월합계에서 제외)")
    f6 = st.file_uploader("선매출", type=['xls','xlsx','csv'], label_visibility="collapsed")
with c5b:
    st.caption("7️⃣ 선발행 (월합계에서 제외)")
    f7 = st.file_uploader("선발행", type=['xls','xlsx','csv'], label_visibility="collapsed")
 
st.markdown("**세금계산서 발행 목록 📋**")
f5 = st.file_uploader("5️⃣ 분기표", type=['xls','xlsx','csv'], label_visibility="collapsed")
 
if f1: st.session_state['매출대장']  = f1.read()
if f2: st.session_state['일반경비']  = f2.read()
if f3: st.session_state['카드매출']  = f3.read()
if f4: st.session_state['현금영수증'] = f4.read()
if f5: st.session_state['계산서']    = f5.read()
if f6: st.session_state['선매출']    = f6.read()
if f7: st.session_state['선발행']    = f7.read()
 
# ── 카드매출 시트 선택 ───────────────────────────────────────
selected_sheet = None
if st.session_state['카드매출']:
    try:
        sheet_names = pd.ExcelFile(BytesIO(st.session_state['카드매출']), engine='openpyxl').sheet_names
        selected_sheet = st.selectbox("📅 카드매출 분석 월 선택", options=sheet_names, index=0)
    except Exception as e:
        st.error(f"카드매출 파일 오류: {e}")
 
# ── 분석 ────────────────────────────────────────────────────
if st.button("🚀 미발행 업체 분석 시작", type="primary", use_container_width=True):
    required = ['매출대장','일반경비','카드매출','현금영수증','계산서']
    missing = [k for k in required if not st.session_state[k]]
    if missing:
        st.warning(f"⚠️ 필수 파일: {', '.join(missing)}"); st.stop()
    if not selected_sheet:
        st.warning("⚠️ 카드매출 월을 선택해주세요."); st.stop()
 
    with st.spinner("분석 중..."):
        try:
            # 1. 매출대장
            df_출고 = read_tsv_cp949(st.session_state['매출대장'])
            df_출고['_key'] = df_출고['거래처'].apply(lambda x: clean(str(x)))
            df_출고['_공급가액'] = df_출고['공급가액+운송'].apply(to_num)
            df_출고['_부가세'] = (df_출고['부가세_품목'].apply(to_num) + df_출고['운송비(부가세)'].apply(to_num))
            df_출고['_합계'] = df_출고['금액'].apply(to_num)
 
            출고_names, original_map = set(), {}
            for v in df_출고['거래처'].dropna():
                v = str(v).strip().strip('"')
                if v and not is_number(v):
                    k = clean(v); 출고_names.add(k); original_map[k] = v
 
            df_출고_v = df_출고[df_출고['_key'].isin(출고_names)]
            출고_공급 = df_출고_v.groupby('_key')['_공급가액'].sum().to_dict()
            출고_부가 = df_출고_v.groupby('_key')['_부가세'].sum().to_dict()
            출고_합계 = df_출고_v.groupby('_key')['_합계'].sum().to_dict()
 
            # 2. 일반경비
            df_경비 = pd.read_excel(BytesIO(st.session_state['일반경비']), header=0, engine='openpyxl')
            df_경비['_key'] = df_경비['거래처명'].apply(lambda x: clean(str(x)))
            df_경비['_공급가액'] = df_경비['공급가액'].apply(to_num)
            df_경비['_부가세'] = df_경비['세액'].apply(to_num)
            df_경비['_합계'] = df_경비['계'].apply(to_num)
 
            경비_names = set()
            for v in df_경비['거래처명'].dropna():
                v = str(v).strip()
                if v and not is_number(v):
                    k = clean(v); 경비_names.add(k)
                    if k not in original_map: original_map[k] = v
 
            df_경비_v = df_경비[df_경비['_key'].isin(경비_names)]
            경비_공급 = df_경비_v.groupby('_key')['_공급가액'].sum().to_dict()
            경비_부가 = df_경비_v.groupby('_key')['_부가세'].sum().to_dict()
            경비_합계 = df_경비_v.groupby('_key')['_합계'].sum().to_dict()
 
            all_target_names = 출고_names | 경비_names
            all_공급 = {k: 출고_공급.get(k,0)+경비_공급.get(k,0) for k in all_target_names}
            all_부가 = {k: 출고_부가.get(k,0)+경비_부가.get(k,0) for k in all_target_names}
            all_합계 = {k: 출고_합계.get(k,0)+경비_합계.get(k,0) for k in all_target_names}
 
            # 3. 카드매출 — 금액 집계
            df_카드 = pd.read_excel(BytesIO(st.session_state['카드매출']),
                                    sheet_name=selected_sheet, header=2, engine='openpyxl')
            df_카드['_key'] = df_카드['거래처'].apply(lambda x: clean(str(x)))
            df_카드['_금액'] = df_카드['총매출금액'].apply(to_num)
            df_카드_v = df_카드[df_카드['_key'] != '']
            card_합계 = df_카드_v.groupby('_key')['_금액'].sum().to_dict()
            card_공급 = {k: round(v/1.1) for k,v in card_합계.items()}
            card_부가 = {k: v - round(v/1.1) for k,v in card_합계.items()}
            card_names = set(card_합계.keys())
 
            # 4. 현금영수증 — 금액 집계
            df_현금 = read_tsv_cp949(st.session_state['현금영수증'])
            def extract_cash(row):
                v = row.get('비고','')
                if pd.notna(v):
                    m = re.search(r'현금영수증/([^(]+)', str(v))
                    if m: return m.group(1).strip()
                return str(row.get('거래처',''))
            df_현금['_key'] = df_현금.apply(lambda r: clean(extract_cash(r)), axis=1)
            df_현금['_금액'] = df_현금['금액'].apply(to_num) if '금액' in df_현금.columns else 0
            cash_합계 = df_현금[df_현금['_key']!=''].groupby('_key')['_금액'].sum().to_dict()
            cash_공급 = {k: round(v/1.1) for k,v in cash_합계.items()}
            cash_부가 = {k: v - round(v/1.1) for k,v in cash_합계.items()}
            cash_names = set(cash_합계.keys())
 
            # 4-2. 선매출/선발행 (있을 때만)
            def load_pre(blob):
                if not blob: return {}, {}, {}, set()
                df = read_pre_tsv(blob)
                df['_key'] = df['거래처명'].apply(lambda x: clean(str(x)))
                df['_공급'] = df['공급가액'].apply(to_num)
                df['_부가'] = df['부가세'].apply(to_num)
                df['_합'] = df['금액'].apply(to_num)
                df = df[df['_key'] != '']
                # original_map 갱신
                for v in df['거래처명'].dropna():
                    k = clean(str(v))
                    if k and k not in original_map: original_map[k] = str(v).strip()
                return (df.groupby('_key')['_공급'].sum().to_dict(),
                        df.groupby('_key')['_부가'].sum().to_dict(),
                        df.groupby('_key')['_합'].sum().to_dict(),
                        set(df['_key'].unique()))
            presale_공급, presale_부가, presale_합계, presale_names = load_pre(st.session_state['선매출'])
            prepub_공급, prepub_부가, prepub_합계, prepub_names = load_pre(st.session_state['선발행'])
 
            # 5. 세금계산서
            tax_b = st.session_state['계산서']
            is_xls = tax_b[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'
            df_tax = pd.read_excel(BytesIO(tax_b), header=None,
                                   engine='xlrd' if is_xls else 'openpyxl')
 
            tax_names, tax_raw_list = set(), []
            for i, v in df_tax.iloc[6:, 11].items():
                if pd.notna(v) and str(v).strip():
                    k = clean(str(v))
                    tax_names.add(k); tax_raw_list.append((i - 5, str(v).strip()))
                    if k not in original_map: original_map[k] = str(v).strip()
 
            df_tax_d = df_tax.iloc[6:].copy()
            df_tax_d['_key'] = df_tax_d[11].apply(lambda x: clean(str(x)))
            df_tax_d['_공급가액'] = df_tax_d[15].apply(to_num)
            df_tax_d['_세액'] = df_tax_d[16].apply(to_num)
            df_tax_d['_합계금액'] = df_tax_d[14].apply(to_num)
            df_tax_v = df_tax_d[df_tax_d['_key'].isin(tax_names)]
            tax_공급 = df_tax_v.groupby('_key')['_공급가액'].sum().to_dict()
            tax_세액 = df_tax_v.groupby('_key')['_세액'].sum().to_dict()
            tax_합계 = df_tax_v.groupby('_key')['_합계금액'].sum().to_dict()
 
            # 6. 거래처별 금액 단위 검증 (set exclude → 금액 차감으로 변경)
            skip_kw = ['기타거래처','현금영수증','카드매출','거래처명','거래처ID','유창']
            analyze_keys = all_target_names | tax_names
 
            missing_rows, amount_rows = [], []
            for k in analyze_keys:
                orig = original_map.get(k, '')
                if not orig: continue
                if any(kw in orig for kw in skip_kw): continue
 
                e_공급 = all_공급.get(k, 0)
                e_부가 = all_부가.get(k, 0)
                e_합계 = all_합계.get(k, 0)
 
                sub_공급 = card_공급.get(k,0)+cash_공급.get(k,0)+presale_공급.get(k,0)+prepub_공급.get(k,0)
                sub_부가 = card_부가.get(k,0)+cash_부가.get(k,0)+presale_부가.get(k,0)+prepub_부가.get(k,0)
                sub_합계 = card_합계.get(k,0)+cash_합계.get(k,0)+presale_합계.get(k,0)+prepub_합계.get(k,0)
 
                adj_공급 = e_공급 - sub_공급
                adj_부가 = e_부가 - sub_부가
                adj_합계 = e_합계 - sub_합계
 
                t_공급 = tax_공급.get(k, 0)
                t_세액 = tax_세액.get(k, 0)
                t_합계 = tax_합계.get(k, 0)
 
                d_공급 = adj_공급 - t_공급
                d_부가 = adj_부가 - t_세액
                d_합계 = adj_합계 - t_합계
 
                # 미발행: 세금계산서 0이고 조정매출 > 1000
                if k not in tax_names and adj_합계 > 1000:
                    similars = find_similar(k, tax_raw_list, 0.72)
                    reason = (" / ".join([f"유사이름: '{s}' (계산서 {r}행, {sc:.0%})"
                                          for s,r,sc in similars])
                              if similars else "세금계산서 미발행")
                    missing_rows.append({"업체명": orig, "조정매출": adj_합계, "비고": reason})
                    continue
 
                # 금액 불일치 (±1,000원 이상)
                if abs(d_합계) > 1000:
                    if d_합계 > 0:
                        상태 = "🚨 누락 (가산세 위험)"
                    elif adj_합계 == 0:
                        상태 = "⚠️ 매출외 발행 (이월/오류)"
                    else:
                        상태 = "⚠️ 과대발행"
                    amount_rows.append({
                        "업체명": orig, "상태": 상태,
                        "ERP_공급가액": e_공급, "ERP_부가세": e_부가, "ERP_합계": e_합계,
                        "차감_합계": sub_합계,
                        "조정매출_공급가액": adj_공급, "조정매출_부가세": adj_부가, "조정매출_합계": adj_합계,
                        "세금계산서_공급가액": t_공급, "세금계산서_세액": t_세액, "세금계산서_합계": t_합계,
                        "차이_공급가액": d_공급, "차이_부가세": d_부가, "차이_합계": d_합계,
                    })
 
            missing_rows.sort(key=lambda x: -x["조정매출"])
            amount_rows.sort(key=lambda x: -abs(x["차이_합계"]))
 
        except Exception as e:
            st.error(f"❌ 오류: {e}")
            import traceback; st.code(traceback.format_exc()); st.stop()
 
    # ── 결과 ─────────────────────────────────────────────────
    st.divider()
    c1,c2,c3,c4,c5,c6 = st.columns(6)
    c1.metric("매출대장", f"{len(출고_names)}개")
    c2.metric("일반경비", f"{len(경비_names)}개")
    c3.metric("카드+현금", f"{len(card_names | cash_names)}개")
    c4.metric("선매출+선발행", f"{len(presale_names | prepub_names)}개")
    c5.metric("세금계산서", f"{len(tax_names)}개")
    c6.metric("⚠️ 미발행", f"{len(missing_rows)}개",
              delta=f"-{len(missing_rows)}", delta_color="inverse")
 
    tab1, tab2 = st.tabs([
        f"📋 미발행 ({len(missing_rows)})",
        f"💰 금액 불일치 ({len(amount_rows)})"
    ])
 
    with tab1:
        st.subheader(f"[{selected_sheet}] 세금계산서 미발행 업체")
        if missing_rows:
            df_m = pd.DataFrame([{
                "No.": i+1, "업체명": r["업체명"],
                "조정매출": fmt(r["조정매출"]), "비고": r["비고"]
            } for i,r in enumerate(missing_rows)])
            def hl_m(row):
                return (['background-color:#fff9c4']*len(row)
                        if "유사이름" in str(row["비고"]) else ['']*len(row))
            st.dataframe(df_m.style.apply(hl_m,axis=1),
                         use_container_width=True, hide_index=True)
            st.caption("🟡 = 세금계산서에 유사이름 있음 (직접 확인) | 💡 조정매출 = 매출+경비−카드−현금−선매출−선발행")
        else:
            st.success("🎉 미발행 업체 없음!")
 
    with tab2:
        st.subheader(f"[{selected_sheet}] 금액 불일치")
        if amount_rows:
            df_a = pd.DataFrame([{
                "No.": i+1, "업체명": r["업체명"], "상태": r["상태"],
                "ERP 합계": fmt(r["ERP_합계"]),
                "차감(카드/현/선)": fmt(r["차감_합계"]),
                "조정매출": fmt(r["조정매출_합계"]),
                "세금계산서": fmt(r["세금계산서_합계"]),
                "차이": fmt(r["차이_합계"]),
            } for i,r in enumerate(amount_rows)])
            col_colors = {
                "No.": "#F5F5F5", "업체명": "#E8F4FD", "상태": "#FFF3E0",
                "ERP 합계": "#FFE0B2", "차감(카드/현/선)": "#FFCCBC",
                "조정매출": "#FFAB91", "세금계산서": "#E1BEE7", "차이": "#EF9A9A",
            }
            def hl_a(row):
                return [f"background-color:{col_colors.get(c,'#FFF')}" for c in row.index]
            st.dataframe(df_a.style.apply(hl_a,axis=1),
                         use_container_width=True, hide_index=True)
            st.caption("🚨 누락 = 조정매출 > 세금계산서 | ⚠️ 과대발행 = 세금계산서 > 조정매출")
        else:
            st.success("🎉 금액 모두 일치!")
 
    # ── 엑셀 다운로드 ────────────────────────────────────────
    wb = Workbook()
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'))
    def mk_hdr(ws, titles, txt):
        ws['A1'] = txt
        ws['A1'].font = Font(bold=True, size=13, name="맑은 고딕")
        ws.merge_cells(f'A1:{chr(64+len(titles))}1')
        for ci, t in enumerate(titles, 1):
            c = ws.cell(2, ci, t)
            c.fill = PatternFill("solid", fgColor="CC2222")
            c.font = Font(bold=True, color="FFFFFF", name="맑은 고딕")
            c.alignment = Alignment(horizontal='center'); c.border = border
 
    ws1 = wb.active; ws1.title = "미발행업체"
    mk_hdr(ws1, ["No.","업체명","조정매출","비고"], f"세금계산서 미발행 ({selected_sheet})")
    yf = PatternFill("solid", fgColor="FFF9C4")
    af = PatternFill("solid", fgColor="FFF0F0")
    for i, r in enumerate(missing_rows, 1):
        ws1.cell(i+2,1,i); ws1.cell(i+2,2,r["업체명"])
        ws1.cell(i+2,3,r["조정매출"]).number_format='#,##0'
        ws1.cell(i+2,4,r["비고"])
        hl = yf if "유사이름" in r["비고"] else (af if i%2==0 else None)
        for ci in range(1,5):
            c = ws1.cell(i+2,ci)
            c.font = Font(name="맑은 고딕", size=10); c.border = border
            c.alignment = Alignment(horizontal='right' if ci==3 else ('left' if ci>1 else 'center'),
                                    vertical='center', wrap_text=(ci==4))
            if hl: c.fill = hl
    for ci, w in enumerate([6,30,14,50], 1): ws1.column_dimensions[chr(64+ci)].width = w
    for r in range(2, len(missing_rows)+3): ws1.row_dimensions[r].height = 26
 
    ws2 = wb.create_sheet("금액불일치")
    h2 = ["No.","업체명","상태","ERP합계","차감(카드/현/선)","조정매출","세금계산서","차이"]
    mk_hdr(ws2, h2, f"금액 불일치 ({selected_sheet})")
    xls_colors = ["F5F5F5","E8F4FD","FFF3E0","FFE0B2","FFCCBC","FFAB91","E1BEE7","EF9A9A"]
    nc = {4,5,6,7,8}
    for i, r in enumerate(amount_rows, 1):
        vals = [i, r["업체명"], r["상태"],
                r["ERP_합계"], r["차감_합계"], r["조정매출_합계"], r["세금계산서_합계"], r["차이_합계"]]
        for ci, val in enumerate(vals, 1):
            c = ws2.cell(i+2, ci, val)
            c.font = Font(name="맑은 고딕", size=10); c.border = border
            c.fill = PatternFill("solid", fgColor=xls_colors[ci-1])
            c.alignment = Alignment(
                horizontal='right' if ci in nc else ('center' if ci==1 else 'left'),
                vertical='center')
            if ci in nc: c.number_format = '#,##0'
    for ci, w in enumerate([6,28,18,14,16,14,14,14], 1): ws2.column_dimensions[chr(64+ci)].width = w
    for r in range(2, len(amount_rows)+3): ws2.row_dimensions[r].height = 22
 
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    st.download_button(
        "⬇️ 전체 결과 엑셀 다운로드",
        data=buf, file_name=f"마감체크_{selected_sheet}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True, type="primary"
    )
 
